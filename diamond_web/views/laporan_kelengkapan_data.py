"""Laporan Kelengkapan Data views - Quality Control Reports."""

import django
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_GET, require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_protect
from datetime import datetime, timedelta
from django.db.models import Q
from io import BytesIO
from openpyxl import Workbook

from ..models import Tiket
from ..constants.tiket_status import STATUS_LABELS
from ..forms.laporan_kelengkapan_data import LaporanKelengkapanDataFilterForm, TiketExportResource

def is_pmde_user(user):
    """Check if user belongs to PMDE group."""
    return user.is_superuser or user.is_staff or user.groups.filter(name__in=['user_pmde', 'admin', 'admin_pmde']).exists()

class LaporanKelengkapanDataView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Quality Control Report with filtering by quarter (triwulan) and year."""
    template_name = 'laporan_kelengkapan_data/list.html'

    def test_func(self):
        """Allow access only to PMDE users."""
        return is_pmde_user(self.request.user)
    
    def get_context_data(self, **kwargs):
        """Add filter form and available years to context."""
        context = super().get_context_data(**kwargs)
        # Get distinct years from Tiket data for filter options
        tikets = Tiket.objects.all()
        years = sorted(set(t.tahun for t in tikets), reverse=True)

        current_year = datetime.now().year
        # If no data, default to current year   
        if current_year not in years:
            years.insert(0, current_year)

        context['years'] = years
        context['form'] = LaporanKelengkapanDataFilterForm(years=years)
        return context


@login_required
@user_passes_test(is_pmde_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_kelengkapan_data_data(request): 
    """DataTables server-side endpoint for Laporan Kelengkapan Data."""
    params = request.POST if request.method == 'POST' else request.GET
    periode_type = params.get('periode_type')
    periode = params.get('periode')
    tahun = params.get('tahun')
    try:
        draw = int(params.get('draw', 1))
    except (ValueError, TypeError):
        draw = 1   

    try:
        start = int(params.get('start', 0))
    except (ValueError, TypeError):
        start = 0

    try:
        length = int(params.get('length', 10))
    except (ValueError, TypeError):
        length = 10

    if not periode_type or not periode or not tahun:
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
        })
    
    try:
        tahun_int = int(tahun)
    except (ValueError, TypeError):
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        })
    
    #Calculate date range based on periode_type and periode
    start_date = None
    end_date = None
    
    if periode_type == 'bulanan':
        try:
            bulan = int(periode)
            if bulan < 1 or bulan > 12:
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })

            start_date = datetime(tahun_int, bulan, 1)
            if bulan == 12:
                end_date = datetime(tahun_int + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(tahun_int, bulan + 1, 1) - timedelta(days=1)    
        except (ValueError, TypeError):
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })

    elif periode_type == 'triwulanan':
        quarter_months = {
            1: (1, 3),  #int(periode)
            2: (4, 6),  #int(periode)
            3: (7, 9),  #int(periode)
            4: (10, 12) #int(periode)
        }
        try:
            triwulan = int(periode)
            if triwulan not in quarter_months:
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered':  0,
                    'data': []
                })
            start_month, end_month = quarter_months[triwulan]
            start_date = datetime(tahun_int, start_month, 1)
            if end_month == 12:
                end_date = datetime(tahun_int + 1, 1, 1 ) - timedelta(days=1)
            else:
                end_date = datetime(tahun_int, end_month + 1, 1) - timedelta(days=1)        
        except (ValueError, TypeError):
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })  
    elif periode_type == 'semester':
        semester_months = {
            1: (1, 6),  #int(periode)
            2: (7, 12) #int(periode)
        }
        try:
            semester = int(periode)
            if semester not in semester_months:
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })
            start_month, end_month = semester_months[semester]
            start_date = datetime(tahun_int, start_month, 1)
            if end_month == 12:
                end_date = datetime(tahun_int + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(tahun_int, end_month + 1, 1) - timedelta(days=1)        
        except (ValueError, TypeError):
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })
    elif periode_type == 'tahunan':
        start_date = datetime(tahun_int, 1, 1)
        end_date = datetime(tahun_int + 1, 1, 1) - timedelta(days=1)
    else:     
        return JsonResponse({
            'draw': draw,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        })
    
    #Query tikets with tgl_transfer in the calculated date range
    qs = Tiket.objects.filter(
        tgl_transfer__isnull=False,
        tgl_transfer__date__gte=start_date.date(),
        tgl_transfer__date__lte=end_date.date()
    ).select_related(
        'id_periode_data',
        'id_periode_data__id_sub_jenis_data_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap', 
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel'
    ).distinct()

    # Search logic
    search_value = params.get('search[value]', '')
    if search_value:
        qs = qs.filter(
            Q(id_periode_data__id_sub_jenis_data_ilap__id_ilap__nama_ilap__icontains=search_value) |
            Q(id_periode_data__id_sub_jenis_data_ilap__nama_sub_jenis_data__icontains=search_value) |
            Q(nomor_tiket__icontains=search_value)
        )

    # Ordering logic
    order_column_idx = params.get('order[0][column]')
    order_dir = params.get('order[0][dir]', 'asc')
    order_columns = [
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap__nama_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__nama_sub_jenis_data',
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel__deskripsi',
        'nomor_tiket',
        'status_tiket',
        'baris_diterima',
        'qc_c'
    ]
    if order_column_idx is not None:
        try:
            idx = int(order_column_idx)
            if idx < len(order_columns):
                order_field = order_columns[idx]
                if order_dir == 'desc':
                    order_field = f'-{order_field}'
                qs = qs.order_by(order_field)
        except (ValueError, TypeError):
            qs = qs.order_by('-tgl_transfer')
    else:
        qs = qs.order_by('-tgl_transfer')

    records_total = Tiket.objects.count()
    records_filtered = qs.count()
    qs_page = qs[start:start + length]
    data = []
    for tiket in qs_page:
        # Safely access related data
        pd = tiket.id_periode_data
        subjenis_data = pd.id_sub_jenis_data_ilap if pd else None
        ilap = subjenis_data.id_ilap if subjenis_data else None
        jenis_tabel = subjenis_data.id_jenis_tabel if subjenis_data else None

        row = {
             'nama_ilap': ilap.nama_ilap if ilap else '',
             'nama_sub_jenis_data': subjenis_data.nama_sub_jenis_data if subjenis_data else '',
             'nama_tabel': jenis_tabel.deskripsi if jenis_tabel else '',
             'nomor_tiket': tiket.nomor_tiket if tiket.nomor_tiket else '',
             'data_diterima': tiket.baris_diterima or 0,
             'status_tiket': STATUS_LABELS.get(tiket.status_tiket, 'Unknown'),
             'qc_c': tiket.qc_c if tiket.qc_c is not None else 0
        }
        data.append(row)
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
@user_passes_test(is_pmde_user)
@require_GET
@csrf_protect
def laporan_kelengkapan_data_export(request):
    """Export Laporan Kelengkapan Data to Excel based on current filters."""
    periode_type = request.GET.get('periode_type')
    periode = request.GET.get('periode')
    tahun = request.GET.get('tahun')

    if not periode_type or not periode or not tahun:
        return JsonResponse({'error': 'Periode type, periode, dan tahun harus dipilih.'}, status=400)

    try:
        tahun= int(tahun)
    except (ValueError, TypeError):
        return HttpResponse('Invalid year', status=400)

    # Calculate date range based on filters (same logic as data endpoint)
    start_date = None
    end_date = None
    periode_label = ''
    
    if periode_type == 'bulanan':
        try:
            bulan = int(periode)
            if bulan < 1 or bulan > 12:
                return JsonResponse({'error': 'Bulan tidak valid.'}, status=400)

            start_date = datetime(tahun, bulan, 1)
            if bulan == 12:
                end_date = datetime(tahun + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(tahun, bulan + 1, 1) - timedelta(days=1)    
            bulan_names = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
            periode_label = f"{bulan_names[bulan]} {tahun}" 

        except (ValueError, TypeError):
            return JsonResponse({'error': 'Bulan tidak valid.'}, status=400)
        
    elif periode_type == 'triwulanan':
        quarter_months = {
            1: (1, 3),
            2: (4, 6),
            3: (7, 9),
            4: (10, 12)
        }
        try:
            triwulan = int(periode)
            if triwulan not in quarter_months:
                return JsonResponse({'error': 'Triwulan tidak valid.'}, status=400)
            start_month, end_month = quarter_months[triwulan]
            start_date = datetime(tahun, start_month, 1)
            if end_month == 12:
                end_date = datetime(tahun + 1, 1, 1 ) - timedelta(days=1)
            else:
                end_date = datetime(tahun, end_month + 1, 1) - timedelta(days=1)        
            periode_label = f"Triwulan_{triwulan}_{tahun}"
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Triwulan tidak valid.'}, status=400) 
    elif periode_type == 'semester':
        semester_months = {
            1: (1, 6),
            2: (7, 12)
        }   
        try:
            semester = int(periode)
            if semester not in semester_months:
                return JsonResponse({'error': 'Semester tidak valid.'}, status=400)
            start_month, end_month = semester_months[semester]
            start_date = datetime(tahun, start_month, 1)
            if end_month == 12:
                end_date = datetime(tahun + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(tahun, end_month + 1, 1) - timedelta(days=1)        
            periode_label = f"Semester_{semester}_{tahun}"
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Semester tidak valid.'}, status=400)
    elif periode_type == 'tahunan':
        start_date = datetime(tahun, 1, 1)
        end_date = datetime(tahun + 1, 1, 1) - timedelta(days=1)
        periode_label = f"Tahun_{tahun}"

    else:
        return JsonResponse({'error': 'Jenis periode tidak valid.'}, status=400)
    
    tikets = Tiket.objects.filter(
        tgl_transfer__isnull=False,
        tgl_transfer__date__gte=start_date.date(),
        tgl_transfer__date__lte=end_date.date()
    ).select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel'
    ).order_by('-tgl_transfer')
    # Create Excel workbook

    resource= TiketExportResource()
    dataset = resource.export(tikets)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tikets"

    # Write header
    headers = dataset.headers
    for col_idx, row in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=row)
    # Write data rows
    for row_idx, row in enumerate(dataset, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    #save to BytesIO
    output = BytesIO()
    wb.save(output)
    excel_data = output.getvalue()
    
    #Create HTTP response with Excel file
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laporan_kelengkapan_data_{periode_label}.xlsx"'
    return response